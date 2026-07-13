# -*- coding: utf-8 -*-
"""Importación automatizada del Set de Pruebas DGII (Excel).

Flujo completo:
1. Lee pestaña ECF (+ InformacionReferencia si existe).
2. Crea facturas con tipo e-CF correcto (E31–E47).
3. Ordena: bases primero, luego E33/E34 vinculadas.
4. Opcionalmente confirma y emite e-CF (bases → notas).
"""
import base64
import io
import logging
import re

import openpyxl
from odoo import models, fields, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Prioridad de creación/emisión: bases antes que notas de débito/crédito
_TIPO_ORDER = {
    31: 10, 32: 10, 41: 20, 43: 20, 44: 20, 45: 20, 46: 20, 47: 20,
    33: 90, 34: 90,
}


def _cell_str(val):
    if val is None or val == '#e' or val is False:
        return False
    if isinstance(val, float) and val == int(val):
        val = int(val)
    return str(val).strip()


def _normalize_rnc(val):
    raw = _cell_str(val)
    if not raw:
        return False
    digits = ''.join(c for c in raw if c.isdigit())
    return digits or False


def _ensure_itbis_tax(env, amount, tax_type, company):
    """Busca o crea ITBIS de venta/compra (16 o 18)."""
    Tax = env['account.tax'].sudo()
    domain = [
        ('amount', '=', float(amount)),
        ('type_tax_use', '=', tax_type),
        ('company_id', '=', company.id),
        ('amount_type', '=', 'percent'),
    ]
    tax = Tax.search(domain, limit=1)
    if tax:
        return tax
    tax = Tax.search([
        ('amount', '=', float(amount)),
        ('type_tax_use', '=', tax_type),
        ('company_id', 'in', [company.id, False]),
    ], limit=1)
    if tax:
        return tax
    return Tax.create({
        'name': f'ITBIS {int(amount)}%',
        'amount': float(amount),
        'amount_type': 'percent',
        'type_tax_use': tax_type,
        'company_id': company.id,
        'description': f'ITBIS {int(amount)}% (homologación DGII)',
    })


def _header_index(headers, *candidates):
    lower = {str(h).strip().lower(): i for i, h in enumerate(headers)}
    for name in candidates:
        if name.lower() in lower:
            return lower[name.lower()]
    for i, h in enumerate(headers):
        hl = str(h).strip().lower()
        for name in candidates:
            if name.lower() in hl:
                return i
    return -1


def _load_referencia_map(wb):
    """ENCF → NCFModificado desde pestaña InformacionReferencia (u homónimas)."""
    mapping = {}
    for sheet_name in wb.sheetnames:
        if not re.search(r'referenc', sheet_name, re.I):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else '' for h in rows[0]]
        encf_i = _header_index(headers, 'ENCF', 'eNCF', 'NCF')
        mod_i = _header_index(headers, 'NCFModificado', 'NCFMod', 'ENCFModificado')
        if encf_i < 0 or mod_i < 0:
            continue
        for r in rows[1:]:
            if not r or not any(r):
                continue
            encf = _cell_str(r[encf_i] if encf_i < len(r) else None)
            mod = _cell_str(r[mod_i] if mod_i < len(r) else None)
            if encf and mod:
                mapping[encf] = mod
    return mapping


class EcfHomologacionWizard(models.TransientModel):
    _name = 'ecf.homologacion.wizard'
    _description = 'Asistente de Homologación DGII'

    xlsx_file = fields.Binary(string='Archivo Excel de Pruebas (.xlsx)', required=True)
    filename = fields.Char(string='Nombre de Archivo')
    auto_confirm = fields.Boolean(
        string='Confirmar facturas',
        default=True,
        help='Pasa las facturas de Borrador a Registrado automáticamente.',
    )
    auto_emit = fields.Boolean(
        string='Emitir e-CF automáticamente',
        default=True,
        help='Emite a la DGII en orden: E31/E32/… primero; luego E33/E34 ya vinculadas.',
    )

    def action_importar_set_pruebas(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError(_("Por favor, sube el archivo Excel descargado desde la DGII."))

        file_data = base64.b64decode(self.xlsx_file)
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_("Error al leer el archivo Excel: %s") % str(e)) from e

        if 'ECF' not in wb.sheetnames:
            raise UserError(_("El archivo subido no contiene la pestaña 'ECF' requerida por la DGII."))

        ref_map = _load_referencia_map(wb)
        ws = wb['ECF']
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("La pestaña 'ECF' está vacía."))

        headers = [str(h) if h is not None else '' for h in rows[0]]
        tipo_idx = _header_index(headers, 'TipoeCF')
        ncf_idx = _header_index(headers, 'ENCF', 'eNCF')
        rnc_comp_idx = _header_index(headers, 'RNCComprador')
        id_ext_idx = _header_index(headers, 'IdentificadorExtranjero')
        rs_comp_idx = _header_index(headers, 'RazonSocialComprador')
        total_idx = _header_index(headers, 'MontoTotal')
        discount_idx = _header_index(headers, 'MontoDescuentooRecargo[1]', 'MontoDescuento')
        ncf_mod_idx = _header_index(headers, 'NCFModificado', 'NCFMod')
        rs_emisor_idx = _header_index(headers, 'RazonSocialEmisor', 'NombreComercial')

        if tipo_idx == -1:
            raise UserError(_("No se encontró la columna 'TipoeCF' en el Excel."))

        parsed = []
        errors = []
        for r_idx, r in enumerate(rows[1:], start=2):
            case = self._parse_excel_row(
                r_idx, r, headers,
                tipo_idx, ncf_idx, rnc_comp_idx, id_ext_idx, rs_comp_idx,
                total_idx, discount_idx, ncf_mod_idx, rs_emisor_idx, ref_map,
            )
            if case is None:
                continue
            if case.get('error'):
                errors.append(case['error'])
                continue
            parsed.append(case)

        parsed.sort(key=lambda c: (_TIPO_ORDER.get(c['tipo'], 50), c['ncf_esperado'] or '', c['r_idx']))

        country_do = self.env['res.country'].search([('code', '=', 'DO')], limit=1)
        cf_partner = self._get_or_create_cf_partner(country_do)

        moves_by_encf = {}
        created_moves = self.env['account.move']
        pending_refs = []

        for case in parsed:
            try:
                move = self._create_move_from_case(case, cf_partner, country_do)
            except Exception as e:
                _logger.exception('Homologación fila %s: %s', case['r_idx'], e)
                errors.append(_('Fila %s (NCF %s): %s') % (
                    case['r_idx'], case['ncf_esperado'] or '—', e,
                ))
                continue

            created_moves |= move
            if case['ncf_esperado']:
                moves_by_encf[case['ncf_esperado']] = move
            if case['tipo'] in (33, 34):
                ncf_mod = case.get('ncf_modificado')
                if ncf_mod:
                    pending_refs.append((move, ncf_mod))
                else:
                    ncf_mod = self._infer_ncf_modificado(case, parsed)
                    if ncf_mod:
                        pending_refs.append((move, ncf_mod))
                        move.message_post(
                            body=_('Vínculo E%s inferido a %s (sin NCFModificado en Excel).') % (
                                case['tipo'], ncf_mod,
                            ),
                            message_type='comment',
                        )
                    else:
                        errors.append(_(
                            'E%s %s: sin NCFModificado; vincule manualmente la factura original.'
                        ) % (case['tipo'], case['ncf_esperado'] or move.id))

        self._apply_references(pending_refs, moves_by_encf, errors)

        confirmed = 0
        emitted = 0
        if self.auto_confirm and created_moves:
            confirmed, conf_errors = self._confirm_moves(created_moves)
            errors.extend(conf_errors)

        if self.auto_emit and created_moves:
            emitted, emit_errors = self._emit_moves_ordered(created_moves)
            errors.extend(emit_errors)

        msg_parts = [
            _('Importación DGII: %s facturas creadas.') % len(created_moves),
        ]
        if self.auto_confirm:
            msg_parts.append(_('%s confirmadas.') % confirmed)
        if self.auto_emit:
            msg_parts.append(_('%s e-CF emitidos.') % emitted)
        if ref_map:
            msg_parts.append(_('%s referencias leídas.') % len(ref_map))
        msg = ' '.join(msg_parts)
        notif_type = 'success'
        if errors:
            msg = '%s\n\n%s' % (msg, '\n'.join(errors[:20]))
            if len(errors) > 20:
                msg += '\n…'
            notif_type = 'warning' if created_moves else 'danger'

        params = {
            'title': _('Set de Pruebas DGII'),
            'message': msg,
            'type': notif_type,
            'sticky': bool(errors),
        }
        if created_moves:
            params['next'] = {
                'type': 'ir.actions.act_window',
                'name': _('Facturas Set DGII'),
                'res_model': 'account.move',
                'view_mode': 'list,form',
                'domain': [('id', 'in', created_moves.ids)],
                'context': {'create': False},
            }
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': params,
        }

    # ── Helpers ────────────────────────────────────────────────────────────

    def _get_or_create_cf_partner(self, country_do):
        Partner = self.env['res.partner']
        partner = Partner.search([
            ('name', '=', 'CONSUMIDOR FINAL DGII'),
            ('vat', 'in', [False, '']),
            ('company_id', 'in', [False, self.env.company.id]),
        ], limit=1)
        if partner:
            return partner
        return Partner.create({
            'name': 'CONSUMIDOR FINAL DGII',
            'vat': False,
            'country_id': country_do.id if country_do else False,
            'company_id': self.env.company.id,
        })

    def _parse_excel_row(
        self, r_idx, r, headers,
        tipo_idx, ncf_idx, rnc_comp_idx, id_ext_idx, rs_comp_idx,
        total_idx, discount_idx, ncf_mod_idx, rs_emisor_idx, ref_map,
    ):
        if not any(r):
            return None
        tipo_raw = r[tipo_idx] if tipo_idx < len(r) else None
        if not tipo_raw or tipo_raw == '#e':
            return None
        try:
            tipo = int(float(tipo_raw))
        except (TypeError, ValueError):
            return {'error': _('Fila %s: TipoeCF inválido (%s)') % (r_idx, tipo_raw)}

        ncf_esperado = _cell_str(r[ncf_idx]) if ncf_idx != -1 and ncf_idx < len(r) else False
        rnc_clean = _normalize_rnc(r[rnc_comp_idx]) if rnc_comp_idx != -1 and rnc_comp_idx < len(r) else False
        id_ext = _cell_str(r[id_ext_idx]) if id_ext_idx != -1 and id_ext_idx < len(r) else False
        rs_comprador = _cell_str(r[rs_comp_idx]) if rs_comp_idx != -1 and rs_comp_idx < len(r) else False
        rs_emisor = _cell_str(r[rs_emisor_idx]) if rs_emisor_idx != -1 and rs_emisor_idx < len(r) else False

        ncf_modificado = False
        if ncf_mod_idx != -1 and ncf_mod_idx < len(r):
            ncf_modificado = _cell_str(r[ncf_mod_idx])
        if not ncf_modificado and ncf_esperado and ncf_esperado in ref_map:
            ncf_modificado = ref_map[ncf_esperado]

        total_val = 0.0
        if total_idx != -1 and total_idx < len(r) and r[total_idx] not in (None, '#e'):
            try:
                total_val = float(r[total_idx])
            except (TypeError, ValueError):
                total_val = 0.0

        discount_val = 0.0
        if discount_idx != -1 and discount_idx < len(r) and r[discount_idx] not in (None, '#e'):
            try:
                discount_val = float(r[discount_idx])
            except (TypeError, ValueError):
                discount_val = 0.0

        is_rfce = (
            tipo == 32
            and total_val > 0
            and total_val < 250000
            and bool(rs_emisor and '250' in rs_emisor.upper())
        )

        return {
            'r_idx': r_idx,
            'row': r,
            'headers': headers,
            'tipo': tipo,
            'ncf_esperado': ncf_esperado,
            'rnc_clean': rnc_clean,
            'id_ext': id_ext,
            'rs_comprador': rs_comprador,
            'ncf_modificado': ncf_modificado,
            'total_val': total_val,
            'discount_val': discount_val,
            'is_rfce': is_rfce,
        }

    def _infer_ncf_modificado(self, case, all_cases):
        """Si falta NCFModificado, toma el primer E31/E32 del set con mismo RNC."""
        rnc = case.get('rnc_clean')
        for other in all_cases:
            if other.get('error') or other['tipo'] not in (31, 32):
                continue
            if not other.get('ncf_esperado'):
                continue
            if rnc and other.get('rnc_clean') and other['rnc_clean'] != rnc:
                continue
            if other['ncf_esperado'] == case.get('ncf_esperado'):
                continue
            return other['ncf_esperado']
        for other in all_cases:
            if other.get('tipo') in (31, 32) and other.get('ncf_esperado'):
                return other['ncf_esperado']
        return False

    def _resolve_partner(self, case, cf_partner, country_do):
        tipo = case['tipo']
        rnc_clean = case['rnc_clean']
        id_ext = case['id_ext']
        rs = case['rs_comprador']
        Partner = self.env['res.partner']

        if rnc_clean:
            if len(rnc_clean) not in (9, 11):
                raise UserError(_('RNC inválido %s') % rnc_clean)
            partner = Partner.search([('vat', '=', rnc_clean)], limit=1)
            if not partner:
                partner = Partner.create({
                    'name': rs or f'DGII Test Cliente {rnc_clean}',
                    'vat': rnc_clean,
                    'country_id': country_do.id if country_do else False,
                })
            return partner

        # E32 consumidor final / E43 gastos menores sin RNC
        if tipo in (32, 43):
            return cf_partner

        # E47 pagos al exterior: identificador extranjero
        if tipo == 47 and id_ext:
            partner = Partner.search([
                '|', ('vat', '=', id_ext), ('ref', '=', id_ext),
            ], limit=1)
            if not partner:
                partner = Partner.create({
                    'name': rs or f'DGII Exterior {id_ext}',
                    'vat': False,
                    'ref': id_ext,
                    'country_id': country_do.id if country_do else False,
                })
            return partner

        if tipo in (31, 33, 34, 41, 44, 45, 46, 47):
            raise UserError(_('E%s requiere RNCComprador') % tipo)

        return cf_partner

    def _build_lines(self, case):
        r = case['row']
        headers = case['headers']
        tipo = case['tipo']
        tax_type = 'sale'
        line_ids = []
        missing_tax = False

        for item_num in range(1, 16):
            desc_col = qty_col = price_col = tax_ind_col = None
            for idx, h in enumerate(headers):
                hs = str(h)
                if f'NombreItem[{item_num}]' in hs:
                    desc_col = idx
                elif f'CantidadItem[{item_num}]' in hs:
                    qty_col = idx
                elif f'PrecioUnitarioItem[{item_num}]' in hs:
                    price_col = idx
                elif f'IndicadorFacturacion[{item_num}]' in hs:
                    tax_ind_col = idx

            if desc_col is None or desc_col >= len(r) or r[desc_col] in (None, '#e'):
                continue

            item_name = str(r[desc_col])
            qty = 1.0
            price = 0.0
            tax_ind = 4
            if qty_col is not None and qty_col < len(r) and r[qty_col] not in (None, '#e'):
                qty = float(r[qty_col])
            if price_col is not None and price_col < len(r) and r[price_col] not in (None, '#e'):
                price = float(r[price_col])
            if tax_ind_col is not None and tax_ind_col < len(r) and r[tax_ind_col] not in (None, '#e'):
                tax_ind = int(float(r[tax_ind_col]))

            tax = False
            if tax_ind == 1:
                tax = _ensure_itbis_tax(self.env, 18.0, tax_type, self.env.company)
            elif tax_ind == 2:
                tax = _ensure_itbis_tax(self.env, 16.0, tax_type, self.env.company)
            if tax_ind in (1, 2) and not tax:
                missing_tax = True

            line_ids.append((0, 0, {
                'name': item_name,
                'quantity': qty,
                'price_unit': price,
                'tax_ids': [(6, 0, [tax.id])] if tax else False,
            }))

        if missing_tax:
            raise UserError(_('No se pudo obtener/crear ITBIS 16%/18%'))

        if case['discount_val'] > 0.0:
            line_ids.append((0, 0, {
                'name': 'Descuento General de Prueba',
                'quantity': 1.0,
                'price_unit': -case['discount_val'],
                'tax_ids': False,
            }))

        total_val = case['total_val']
        if not line_ids:
            if total_val <= 0 and tipo == 34:
                line_ids.append((0, 0, {
                    'name': f'Ajuste NC DGII {case["ncf_esperado"] or ""}',
                    'quantity': 1.0,
                    'price_unit': 0.01,
                    'tax_ids': False,
                }))
            elif total_val <= 0:
                raise UserError(_('Sin ítems y MontoTotal vacío/cero'))
            else:
                line_ids.append((0, 0, {
                    'name': f'Servicio Homologación DGII {case["ncf_esperado"] or ""}',
                    'quantity': 1.0,
                    'price_unit': total_val,
                    'tax_ids': False,
                }))
        else:
            positive = sum(
                float(cmd[2].get('quantity') or 0) * float(cmd[2].get('price_unit') or 0)
                for cmd in line_ids
                if cmd[0] == 0 and float(cmd[2].get('price_unit') or 0) > 0
            )
            if positive <= 0:
                if total_val > 0:
                    line_ids = [(0, 0, {
                        'name': f'Servicio Homologación DGII {case["ncf_esperado"] or ""}',
                        'quantity': 1.0,
                        'price_unit': total_val,
                        'tax_ids': False,
                    })]
                elif tipo == 34:
                    line_ids = [(0, 0, {
                        'name': f'Ajuste NC DGII {case["ncf_esperado"] or ""}',
                        'quantity': 1.0,
                        'price_unit': 0.01,
                        'tax_ids': False,
                    })]
                else:
                    raise UserError(_('Monto resultante RD$0'))

        return line_ids

    def _create_move_from_case(self, case, cf_partner, country_do):
        partner = self._resolve_partner(case, cf_partner, country_do)
        ecf_tipo = self.env['ecf.tipo'].search([('codigo', '=', case['tipo'])], limit=1)
        if not ecf_tipo:
            raise UserError(_('Tipo e-CF %s no existe en el catálogo') % case['tipo'])

        line_ids = self._build_lines(case)
        move_type = 'out_refund' if case['tipo'] == 34 else 'out_invoice'
        ncf_esperado = case['ncf_esperado']
        ref = f'Caso DGII {ncf_esperado}' if ncf_esperado else f'Caso DGII fila {case["r_idx"]}'
        if case.get('is_rfce'):
            ref = f'Caso DGII RFCE {ncf_esperado}'

        vals = {
            'move_type': move_type,
            'partner_id': partner.id,
            'invoice_date': fields.Date.context_today(self),
            'ecf_tipo_id': ecf_tipo.id,
            'ecf_modo': 'inmediato',
            'invoice_line_ids': line_ids,
            'ref': ref,
            'narration': _(
                'Set pruebas DGII · Tipo E%(tipo)s · ENCF esperado %(ncf)s',
                tipo=case['tipo'],
                ncf=ncf_esperado or '—',
            ),
        }
        return self.env['account.move'].create(vals)

    def _apply_references(self, pending_refs, moves_by_encf, errors):
        for move, ncf_mod in pending_refs:
            orig = moves_by_encf.get(ncf_mod)
            if not orig:
                orig = self.env['account.move'].search([
                    ('ref', 'in', [f'Caso DGII {ncf_mod}', f'Caso DGII RFCE {ncf_mod}']),
                    ('company_id', '=', self.env.company.id),
                ], limit=1)
            if orig:
                move.write({'reversed_entry_id': orig.id})
                move.message_post(
                    body=_('Vinculada a factura original %s (%s)') % (
                        orig.name or orig.id, ncf_mod,
                    ),
                    message_type='comment',
                )
            else:
                errors.append(_(
                    'E%s %s: no se encontró original NCFModificado=%s'
                ) % (move.ecf_tipo_id.codigo, move.ref or move.id, ncf_mod))

    def _confirm_moves(self, moves):
        confirmed = 0
        errors = []
        for move in moves.sorted(key=lambda m: _TIPO_ORDER.get(m.ecf_tipo_id.codigo or 99, 50)):
            if move.state != 'draft':
                continue
            try:
                move.action_post()
                confirmed += 1
            except Exception as e:
                _logger.exception('Confirmar %s: %s', move.ref, e)
                errors.append(_('Confirmar %s: %s') % (move.ref or move.name, e))
        return confirmed, errors

    def _emit_moves_ordered(self, moves):
        """Emite bases primero (asignan NCF); luego E33/E34 con ncf_referencia."""
        errors = []
        emitted = 0
        bases = moves.filtered(lambda m: m.ecf_tipo_id and m.ecf_tipo_id.codigo not in (33, 34))
        notes = moves.filtered(lambda m: m.ecf_tipo_id and m.ecf_tipo_id.codigo in (33, 34))

        for move in bases.sorted(key=lambda m: (m.ecf_tipo_id.codigo or 99, m.id)):
            ok, err = self._safe_emit(move)
            if ok:
                emitted += 1
            elif err:
                errors.append(err)

        # Refrescar NCF de originales antes de emitir notas
        if notes:
            notes.mapped('reversed_entry_id').invalidate_recordset(['ecf_ncf', 'ecf_estado'])

        for move in notes.sorted(key=lambda m: m.id):
            ok, err = self._safe_emit(move)
            if ok:
                emitted += 1
            elif err:
                errors.append(err)

        return emitted, errors

    def _safe_emit(self, move):
        if move.state != 'posted':
            return False, _('Emitir %s: no está confirmada') % (move.ref or move.name)
        if move.ecf_ncf and move.ecf_estado not in ('rechazado', 'anulado', 'error', False, None):
            return False, None
        if move.ecf_modo == 'excento':
            return False, None
        try:
            move._validar_pre_emision()
            move._emitir_ecf()
            return True, None
        except Exception as e:
            _logger.exception('Emitir %s: %s', move.ref, e)
            move.message_post(
                body=_('⚠️ Homologación auto-emisión: %s') % e,
                message_type='comment',
            )
            return False, _('Emitir %s: %s') % (move.ref or move.name, e)
