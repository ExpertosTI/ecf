import base64
import io
import logging

import openpyxl
from odoo import models, fields, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


def _cell_str(val):
    if val is None or val == '#e' or val is False:
        return False
    if isinstance(val, float) and val == int(val):
        val = int(val)
    return str(val).strip()


def _normalize_rnc(val):
    raw = _cell_str(val)
    if not raw:
        return False
    digits = ''.join(c for c in raw if c.isdigit())
    return digits or False


def _ensure_itbis_tax(env, amount, tax_type, company):
    """Busca o crea ITBIS de venta/compra (16 o 18)."""
    Tax = env['account.tax'].sudo()
    domain = [
        ('amount', '=', float(amount)),
        ('type_tax_use', '=', tax_type),
        ('company_id', '=', company.id),
        ('amount_type', '=', 'percent'),
    ]
    tax = Tax.search(domain, limit=1)
    if tax:
        return tax
    # Fallback: impuestos compartidos / nombre ITBIS
    tax = Tax.search([
        ('amount', '=', float(amount)),
        ('type_tax_use', '=', tax_type),
        ('company_id', 'in', [company.id, False]),
    ], limit=1)
    if tax:
        return tax
    return Tax.create({
        'name': f'ITBIS {int(amount)}%',
        'amount': float(amount),
        'amount_type': 'percent',
        'type_tax_use': tax_type,
        'company_id': company.id,
        'description': f'ITBIS {int(amount)}% (homologación DGII)',
    })


class EcfHomologacionWizard(models.TransientModel):
    _name = 'ecf.homologacion.wizard'
    _description = 'Asistente de Homologación DGII'

    xlsx_file = fields.Binary(string='Archivo Excel de Pruebas (.xlsx)', required=True)
    filename = fields.Char(string='Nombre de Archivo')

    def action_importar_set_pruebas(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError(_("Por favor, sube el archivo Excel descargado desde la DGII."))

        file_data = base64.b64decode(self.xlsx_file)
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_("Error al leer el archivo Excel: %s") % str(e)) from e

        if 'ECF' not in wb.sheetnames:
            raise UserError(_("El archivo subido no contiene la pestaña 'ECF' requerida por la DGII."))

        ws = wb['ECF']
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("La pestaña 'ECF' está vacía."))

        headers = [str(h) for h in rows[0]]
        tipo_idx = headers.index('TipoeCF') if 'TipoeCF' in headers else -1
        ncf_idx = headers.index('ENCF') if 'ENCF' in headers else -1
        rnc_comp_idx = headers.index('RNCComprador') if 'RNCComprador' in headers else -1
        rs_comp_idx = headers.index('RazonSocialComprador') if 'RazonSocialComprador' in headers else -1
        total_idx = headers.index('MontoTotal') if 'MontoTotal' in headers else -1
        discount_idx = headers.index('MontoDescuentooRecargo[1]') if 'MontoDescuentooRecargo[1]' in headers else -1
        ncf_mod_idx = headers.index('NCFModificado') if 'NCFModificado' in headers else -1

        if tipo_idx == -1:
            raise UserError(_("No se encontró la columna 'TipoeCF' en el Excel."))

        invoices_created = 0
        errors = []
        moves_by_encf = {}
        pending_refs = []  # (move, ncf_modificado)

        country_do = self.env['res.country'].search([('code', '=', 'DO')], limit=1)
        cf_partner = self.env['res.partner'].search([
            ('name', '=', 'CONSUMIDOR FINAL DGII'),
            ('vat', 'in', [False, '']),
            ('company_id', 'in', [False, self.env.company.id]),
        ], limit=1)
        if not cf_partner:
            cf_partner = self.env['res.partner'].create({
                'name': 'CONSUMIDOR FINAL DGII',
                'vat': False,
                'country_id': country_do.id if country_do else False,
                'company_id': self.env.company.id,
            })

        for r_idx, r in enumerate(rows[1:], start=2):
            if not any(r):
                continue

            tipo_raw = r[tipo_idx]
            if not tipo_raw or tipo_raw == '#e':
                continue
            try:
                tipo = int(float(tipo_raw))
            except (TypeError, ValueError):
                errors.append(_('Fila %s: TipoeCF inválido (%s)') % (r_idx, tipo_raw))
                continue

            ncf_esperado = _cell_str(r[ncf_idx]) if ncf_idx != -1 else False
            rnc_clean = _normalize_rnc(r[rnc_comp_idx]) if rnc_comp_idx != -1 else False
            rs_comprador = _cell_str(r[rs_comp_idx]) if rs_comp_idx != -1 else False
            ncf_modificado = _cell_str(r[ncf_mod_idx]) if ncf_mod_idx != -1 else False

            partner = False
            if rnc_clean:
                if len(rnc_clean) not in (9, 11):
                    errors.append(_('Fila %s: RNC inválido %s') % (r_idx, rnc_clean))
                    continue
                partner = self.env['res.partner'].search([('vat', '=', rnc_clean)], limit=1)
                if not partner:
                    partner = self.env['res.partner'].create({
                        'name': rs_comprador or f'DGII Test Cliente {rnc_clean}',
                        'vat': rnc_clean,
                        'country_id': country_do.id if country_do else False,
                    })
            elif tipo == 32:
                partner = cf_partner
            elif tipo in (31, 33, 34, 41, 43, 44, 45, 46, 47):
                errors.append(_('Fila %s (E%s): requiere RNCComprador') % (r_idx, tipo))
                continue
            else:
                partner = cf_partner

            if tipo == 34:
                move_type = 'out_refund'
            elif tipo in (41, 43, 47):
                # Emisión homologación: tratar como venta electrónica hacia DGII
                move_type = 'out_invoice'
            else:
                move_type = 'out_invoice'

            ecf_tipo = self.env['ecf.tipo'].search([('codigo', '=', tipo)], limit=1)
            if not ecf_tipo:
                errors.append(_('Fila %s: tipo e-CF %s no existe en el catálogo') % (r_idx, tipo))
                continue

            line_ids = []
            tax_type = 'sale' if move_type in ('out_invoice', 'out_refund') else 'purchase'
            missing_tax = False

            for item_num in range(1, 16):
                desc_col = qty_col = price_col = tax_ind_col = None
                for idx, h in enumerate(headers):
                    if f'NombreItem[{item_num}]' in h:
                        desc_col = idx
                    elif f'CantidadItem[{item_num}]' in h:
                        qty_col = idx
                    elif f'PrecioUnitarioItem[{item_num}]' in h:
                        price_col = idx
                    elif f'IndicadorFacturacion[{item_num}]' in h:
                        tax_ind_col = idx

                if desc_col is None or r[desc_col] is None or r[desc_col] == '#e':
                    continue

                item_name = str(r[desc_col])
                qty = float(r[qty_col]) if qty_col is not None and r[qty_col] not in (None, '#e') else 1.0
                price = float(r[price_col]) if price_col is not None and r[price_col] not in (None, '#e') else 0.0
                tax_ind = int(float(r[tax_ind_col])) if tax_ind_col is not None and r[tax_ind_col] not in (None, '#e') else 4

                tax = False
                if tax_ind == 1:
                    tax = _ensure_itbis_tax(self.env, 18.0, tax_type, self.env.company)
                elif tax_ind == 2:
                    tax = _ensure_itbis_tax(self.env, 16.0, tax_type, self.env.company)
                if tax_ind in (1, 2) and not tax:
                    missing_tax = True

                line_ids.append((0, 0, {
                    'name': item_name,
                    'quantity': qty,
                    'price_unit': price,
                    'tax_ids': [(6, 0, [tax.id])] if tax else False,
                }))

            if missing_tax:
                errors.append(_(
                    'Fila %s: no se pudo obtener/crear impuesto ITBIS 16%%/18%%. '
                    'Revise Contabilidad → Impuestos.'
                ) % r_idx)
                continue

            discount_val = 0.0
            if discount_idx != -1 and r[discount_idx] not in (None, '#e'):
                discount_val = float(r[discount_idx])
            if discount_val > 0.0:
                line_ids.append((0, 0, {
                    'name': 'Descuento General de Prueba',
                    'quantity': 1.0,
                    'price_unit': -discount_val,
                    'tax_ids': False,
                }))

            total_val = 0.0
            if total_idx != -1 and r[total_idx] not in (None, '#e'):
                total_val = float(r[total_idx])

            if not line_ids:
                if total_val <= 0:
                    errors.append(_('Fila %s: sin ítems y MontoTotal vacío/cero') % r_idx)
                    continue
                line_ids.append((0, 0, {
                    'name': f'Servicio de Homologación DGII {ncf_esperado or ""}',
                    'quantity': 1.0,
                    'price_unit': total_val,
                    'tax_ids': False,
                }))
            else:
                positive = sum(
                    float(cmd[2].get('quantity') or 0) * float(cmd[2].get('price_unit') or 0)
                    for cmd in line_ids
                    if cmd[0] == 0 and float(cmd[2].get('price_unit') or 0) > 0
                )
                if positive <= 0:
                    if total_val <= 0:
                        errors.append(_('Fila %s: monto resultante RD$0') % r_idx)
                        continue
                    line_ids = [(0, 0, {
                        'name': f'Servicio de Homologación DGII {ncf_esperado or ""}',
                        'quantity': 1.0,
                        'price_unit': total_val,
                        'tax_ids': False,
                    })]

            invoice_vals = {
                'move_type': move_type,
                'partner_id': partner.id,
                'invoice_date': fields.Date.context_today(self),
                'ecf_tipo_id': ecf_tipo.id,
                'invoice_line_ids': line_ids,
                'ref': f'Caso DGII {ncf_esperado}' if ncf_esperado else f'Caso DGII fila {r_idx}',
            }

            try:
                move = self.env['account.move'].create(invoice_vals)
                invoices_created += 1
                if ncf_esperado:
                    moves_by_encf[ncf_esperado] = move
                if tipo in (33, 34) and ncf_modificado:
                    pending_refs.append((move, ncf_modificado))
            except Exception as e:
                _logger.exception('Homologación fila %s: %s', r_idx, e)
                errors.append(_('Fila %s (NCF %s): %s') % (r_idx, ncf_esperado or '—', e))

        for move, ncf_mod in pending_refs:
            orig = moves_by_encf.get(ncf_mod)
            if not orig:
                orig = self.env['account.move'].search([
                    ('ref', '=', f'Caso DGII {ncf_mod}'),
                    ('company_id', '=', self.env.company.id),
                ], limit=1)
            if orig:
                move.write({'reversed_entry_id': orig.id})
            else:
                errors.append(_(
                    'E%s %s: no se encontró factura original NCFModificado=%s '
                    '(emite primero el e-CF original y vincúlala).'
                ) % (move.ecf_tipo_id.codigo, move.ref or move.id, ncf_mod))

        msg = _('Se crearon %s facturas de prueba en borrador.') % invoices_created
        notif_type = 'success'
        if errors:
            msg = '%s\n\n%s' % (msg, '\n'.join(errors[:15]))
            if len(errors) > 15:
                msg += '\n…'
            notif_type = 'warning' if invoices_created else 'danger'

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación Set DGII'),
                'message': msg,
                'type': notif_type,
                'sticky': bool(errors),
            },
        }
