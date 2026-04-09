# api_gateway/reportes.py — Generación de reportes DGII 606/607/608
# Formatos: JSON, TXT (formato DGII), Excel (.xlsx), PDF

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Optional

import asyncpg
from fastapi import APIRouter, Depends, Header, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

# ── Constantes DGII ──

TIPOS_ECF = {
    31: "Crédito Fiscal",
    32: "Consumo",
    33: "Nota de Débito",
    34: "Nota de Crédito",
    41: "Compras",
    43: "Gastos Menores",
    44: "Reg. Especiales",
    45: "Gubernamental",
    46: "Exportaciones",
    47: "Pagos Exterior",
}

TIPO_ANULACION = {
    "01": "Deterioro de factura pre-impresa",
    "02": "Errores de impresión (factura pre-impresa)",
    "03": "Impresión defectuosa",
    "04": "Corrección de información",
    "05": "Cambio de productos",
    "06": "Devolución de productos",
    "07": "Omisión de productos",
    "08": "Errores en secuencia de NCF",
    "09": "Por cese de operaciones",
}


class ExportFormat(str, Enum):
    json = "json"
    txt = "txt"
    xlsx = "xlsx"
    pdf = "pdf"


# ── Helpers ──

def _fmt_monto(val) -> str:
    """Formatea monto para TXT DGII: sin comas, 2 decimales."""
    if val is None:
        return "0.00"
    return f"{Decimal(str(val)):.2f}"


def _fmt_str(val) -> str:
    """Convierte valor a string seguro para TXT (None -> vacío), removiendo | y saltos de línea."""
    if val is None:
        return ""
    return str(val).replace("|", " ").replace("\n", " ").replace("\r", "")


def _fmt_fecha(val) -> str:
    """Fecha YYYYMMDD para TXT DGII."""
    if val is None:
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y%m%d")
    return str(val).replace("-", "")[:8]


def _fmt_fecha_display(val) -> str:
    """Fecha legible para Excel/PDF."""
    if val is None:
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%d/%m/%Y")
    return str(val)[:10]


# ═══════════════════════════════════════════
# REPORTE 606 — Compras
# ═══════════════════════════════════════════

HEADERS_606 = [
    "NCF", "RNC Proveedor", "Proveedor", "Tipo Bienes",
    "Fecha Comprobante", "Fecha Pago", "Monto Servicios", "Monto Bienes",
    "Total", "ITBIS Facturado", "ITBIS Retenido", "ISR Retención",
]


def _606_to_txt(rows: list[dict], rnc: str, periodo: str) -> str:
    """Genera TXT formato DGII para reporte 606."""
    lines = []
    # Header: 606|RNC|Periodo|Cantidad
    lines.append(f"606|{rnc}|{periodo.replace('-', '')}|{len(rows)}")
    for r in rows:
        line = "|".join([
            _fmt_str(r.get("ncf")),
            _fmt_str(r.get("rnc_proveedor")),
            _fmt_fecha(r.get("fecha_comprobante")),
            _fmt_fecha(r.get("fecha_pago")),
            _fmt_str(r.get("tipo_bienes")),
            _fmt_monto(r.get("monto_servicios")),
            _fmt_monto(r.get("monto_bienes")),
            _fmt_monto(r.get("total_monto")),
            _fmt_monto(r.get("itbis_facturado")),
            _fmt_monto(r.get("itbis_retenido")),
            _fmt_monto(r.get("isr_retencion")),
        ])
        lines.append(line)
    return "\r\n".join(lines)


# ═══════════════════════════════════════════
# REPORTE 607 — Ventas
# ═══════════════════════════════════════════

HEADERS_607 = [
    "NCF", "Tipo e-CF", "RNC Comprador", "Comprador",
    "Tipo RNC", "Fecha Emisión", "Tipo Ingresos",
    "Monto Facturado", "ITBIS Facturado", "Total", "Tipo Pago",
    "NCF Referencia", "Estado",
]


def _607_to_txt(rows: list[dict], rnc: str, periodo: str) -> str:
    """Genera TXT formato DGII para reporte 607."""
    lines = []
    lines.append(f"607|{rnc}|{periodo.replace('-', '')}|{len(rows)}")
    for r in rows:
        line = "|".join([
            _fmt_str(r.get("ncf")),
            _fmt_str(r.get("tipo_ecf")),
            _fmt_str(r.get("rnc_comprador")),
            _fmt_str(r.get("tipo_rnc_comprador")),
            _fmt_fecha(r.get("fecha_emision")),
            _fmt_str(r.get("tipo_ingresos")),
            _fmt_monto(r.get("monto_facturado")),
            _fmt_monto(r.get("itbis_facturado")),
            _fmt_monto(r.get("total")),
            _fmt_str(r.get("tipo_pago")),
            _fmt_str(r.get("referencia_ncf")),
        ])
        lines.append(line)
    return "\r\n".join(lines)


# ═══════════════════════════════════════════
# REPORTE 608 — Anulaciones
# ═══════════════════════════════════════════

HEADERS_608 = [
    "NCF", "Tipo e-CF", "Fecha Emisión", "Tipo Anulación", "Fecha Anulación",
]


def _608_to_txt(rows: list[dict], rnc: str, periodo: str) -> str:
    """Genera TXT formato DGII para reporte 608."""
    lines = []
    lines.append(f"608|{rnc}|{periodo.replace('-', '')}|{len(rows)}")
    for r in rows:
        line = "|".join([
            _fmt_str(r.get("ncf")),
            _fmt_str(r.get("tipo_ecf")),
            _fmt_fecha(r.get("fecha_emision")),
            _fmt_str(r.get("tipo_anulacion")),
            _fmt_fecha(r.get("fecha_anulacion")),
        ])
        lines.append(line)
    return "\r\n".join(lines)


# ═══════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════

def _to_xlsx(rows: list[dict], headers: list[str], keys: list[str],
             titulo: str, rnc: str, periodo: str) -> bytes:
    """Genera Excel .xlsx con formato profesional."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # ── Estilos ──
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="0A3D62")
    subtitle_font = Font(name="Calibri", size=10, color="666666")
    money_fmt = '#,##0.00'
    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
    )
    alt_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    # ── Título ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, f"Reporte {titulo}").font = title_font

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1, f"RNC: {rnc}  |  Período: {periodo}  |  Registros: {len(rows)}").font = subtitle_font

    # ── Header row ──
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Detect money columns ──
    money_keys = {k for k in keys if any(w in k for w in
                  ("monto", "total", "itbis", "isr", "subtotal", "facturado", "retenido", "retencion"))}

    # ── Data rows ──
    for row_idx, r in enumerate(rows):
        excel_row = header_row + 1 + row_idx
        for col, key in enumerate(keys, 1):
            val = r.get(key)
            # Format dates
            if isinstance(val, (date, datetime)):
                val = _fmt_fecha_display(val)
            elif val is None:
                val = ""
            cell = ws.cell(excel_row, col, val)
            cell.border = thin_border
            if key in money_keys and val != "":
                try:
                    cell.value = float(Decimal(str(val)))
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    pass
            if row_idx % 2 == 1:
                cell.fill = alt_fill

    # ── Auto-width ──
    for col in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(header_row, col).value))
        for row in range(header_row + 1, header_row + 1 + min(len(rows), 50)):
            val = ws.cell(row, col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # ── Freeze header ──
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════

def _to_pdf(rows: list[dict], headers: list[str], keys: list[str],
            titulo: str, rnc: str, periodo: str) -> bytes:
    """Genera PDF con tabla de reporte."""
    from fpdf import FPDF

    # Sanitize text for Helvetica (latin-1 only)
    def _sanitize(text: str) -> str:
        return text.replace("\u2014", "-").replace("\u2013", "-").replace("\u201c", '"').replace("\u201d", '"').replace("\n", " ").replace("\r", "")

    _safe_titulo = _sanitize(titulo)
    _safe_rnc = _sanitize(rnc)
    _safe_periodo = _sanitize(periodo)

    class ReportPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 14)
            self.set_text_color(10, 61, 98)
            self.cell(0, 10, f"Reporte {_safe_titulo}", new_x="LMARGIN", new_y="NEXT", align="C")
            self.set_font("Helvetica", "", 9)
            self.set_text_color(100, 100, 100)
            self.cell(0, 6, f"RNC: {_safe_rnc}  |  Periodo: {_safe_periodo}  |  Registros: {len(rows)}",
                      new_x="LMARGIN", new_y="NEXT", align="C")
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Generado por Renace ECF - {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Pag {self.page_no()}/{{nb}}",
                      align="C")

    pdf = ReportPDF(orientation="L", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # ── Calculate column widths ──
    avail_w = pdf.w - pdf.l_margin - pdf.r_margin
    n_cols = len(headers)
    col_widths = []

    # Use header length as minimum, scale proportionally
    money_keys_set = {k for k in keys if any(w in k for w in
                      ("monto", "total", "itbis", "isr", "subtotal", "facturado", "retenido", "retencion"))}
    for i, h in enumerate(headers):
        base = max(len(h) * 2.2, 18)
        if keys[i] in money_keys_set:
            base = max(base, 28)
        col_widths.append(base)

    # Scale to fit
    total_w = sum(col_widths)
    col_widths = [w * avail_w / total_w for w in col_widths]

    # ── Table header ──
    pdf.set_fill_color(10, 61, 98)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    # ── Table rows ──
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(30, 30, 30)
    for row_idx, r in enumerate(rows):
        if pdf.get_y() > pdf.h - 25:
            pdf.add_page()
            # Reprint header
            pdf.set_fill_color(10, 61, 98)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 7, h, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(30, 30, 30)

        # Alternate row colors
        if row_idx % 2 == 1:
            pdf.set_fill_color(245, 247, 252)
            fill = True
        else:
            pdf.set_fill_color(255, 255, 255)
            fill = True

        for i, key in enumerate(keys):
            val = r.get(key)
            if isinstance(val, (date, datetime)):
                val = _fmt_fecha_display(val)
            elif val is None:
                val = ""
            elif key in money_keys_set:
                val = _fmt_monto(val)
            else:
                val = str(val)
            val = _sanitize(val)

            align = "R" if key in money_keys_set else "L"
            pdf.cell(col_widths[i], 6, val[:50], border=0, fill=fill, align=align)
        pdf.ln()

    # ── Totals row for money columns ──
    if rows and money_keys_set:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 240, 250)
        for i, key in enumerate(keys):
            if key in money_keys_set:
                total = sum(Decimal(str(r.get(key, 0) or 0)) for r in rows)
                pdf.cell(col_widths[i], 7, _fmt_monto(total), border=1, fill=True, align="R")
            elif i == 0:
                pdf.cell(col_widths[i], 7, "TOTALES", border=1, fill=True, align="L")
            else:
                pdf.cell(col_widths[i], 7, "", border=1, fill=True)
        pdf.ln()

    return bytes(pdf.output())


# ═══════════════════════════════════════════
# RESPONSE BUILDERS
# ═══════════════════════════════════════════

def _build_response(
    rows: list[dict],
    fmt: ExportFormat,
    report_num: str,
    headers: list[str],
    keys: list[str],
    titulo: str,
    rnc: str,
    periodo: str,
    txt_builder,
) -> Response:
    """Builds response in the requested format."""

    if fmt == ExportFormat.json:
        return {"periodo": periodo, "registros": rows}

    filename_base = f"{report_num}_{rnc}_{periodo.replace('-', '')}"

    if fmt == ExportFormat.txt:
        content = txt_builder(rows, rnc, periodo)
        return Response(
            content=content.encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.txt"},
        )

    if fmt == ExportFormat.xlsx:
        content = _to_xlsx(rows, headers, keys, titulo, rnc, periodo)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"},
        )

    if fmt == ExportFormat.pdf:
        content = _to_pdf(rows, headers, keys, titulo, rnc, periodo)
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.pdf"},
        )
